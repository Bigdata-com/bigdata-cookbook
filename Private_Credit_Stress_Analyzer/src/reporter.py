from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config.topics import TOPICS
from src.search import SEARCH_DATE_LABEL
from src.utils import console, setup_logger

logger = setup_logger(__name__)

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"
SCORES_CSV = OUTPUT_DIR / "scores.csv"
EXCEL_OUTPUT = OUTPUT_DIR / "private_credit_stress.xlsx"
HTML_OUTPUT = OUTPUT_DIR / "dashboard.html"

HEADER_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
HEADER_FONT = Font(bold=True, color="000000", size=11)
BODY_FONT = Font(color="000000", size=10)
BODY_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


def _topic_display_name(topic_key: str) -> str:
    """Human-readable topic name for Excel headers (e.g. lender_spread_power -> Lender Spread Power)."""
    return topic_key.replace("_", " ").title()


def _layer_display_name(layer: str) -> str:
    """Display layer as Lender, Borrower or Bank."""
    return layer.title() if layer else layer


def _style_header(ws: Any) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _style_body(ws: Any, n_rows: int, n_cols: int) -> None:
    for row_idx in range(2, n_rows + 2):
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = BODY_FILL
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center")


def _auto_width(ws: Any) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def generate_excel(df: pd.DataFrame) -> Path:
    """Generate the multi-sheet Excel workbook from scored DataFrame."""
    wb = Workbook()

    # ── Sheet 1: Lender Ranking ──────────────────────────────────────────────
    ws_lender = wb.active
    ws_lender.title = "Lender Ranking"
    lender_df = (
        df[df["layer"] == "lender"]
        .sort_values("terms_power_score", ascending=False)
        .reset_index(drop=True)
    )
    lender_cols = [
        "entity_name", "ticker", "terms_power_score",
        "positive_count", "negative_count",
        "top_negative_topic", "top_positive_topic",
    ]
    _write_ranked_sheet(ws_lender, lender_df, lender_cols, score_col="terms_power_score")

    # ── Sheet 2: Borrower Distress ───────────────────────────────────────────
    ws_borrower = wb.create_sheet("Borrower Distress")
    borrower_df = (
        df[df["layer"] == "borrower"]
        .sort_values("stress_score", ascending=False)
        .reset_index(drop=True)
    )
    borrower_topic_cols = [
        t["topic_name"]
        for t in TOPICS
        if "borrower" in t["applies_to"] and t["polarity"] == "negative"
    ]
    borrower_cols = ["entity_name", "ticker", "stress_score"] + [
        c for c in borrower_topic_cols if c in borrower_df.columns
    ]
    _write_ranked_sheet(ws_borrower, borrower_df, borrower_cols, score_col="stress_score")

    # ── Sheet 3: Bank Contagion ──────────────────────────────────────────────
    ws_bank = wb.create_sheet("Bank Contagion")
    bank_df = df[df["layer"] == "bank"].copy()
    _gain = bank_df["bank_market_share_gain"].fillna(0) if "bank_market_share_gain" in bank_df.columns else 0
    _pullback = bank_df["bank_credit_pullback"].fillna(0) if "bank_credit_pullback" in bank_df.columns else 0
    bank_df["net_position_score"] = _gain - _pullback
    bank_df = bank_df.sort_values("net_position_score", ascending=False).reset_index(
        drop=True
    )
    bank_topic_cols = [t["topic_name"] for t in TOPICS if "bank" in t["applies_to"]]
    bank_cols = ["entity_name", "ticker", "net_position_score"] + [
        c for c in bank_topic_cols if c in bank_df.columns
    ]
    _write_ranked_sheet(ws_bank, bank_df, bank_cols, score_col="net_position_score")

    # ── Sheets 4–6: Raw Signal Matrix (split by layer) ─────────────────────────
    for layer_key, layer_label in [("lender", "Lenders"), ("borrower", "Borrowers"), ("bank", "Banks")]:
        layer_topics = [str(t["topic_name"]) for t in TOPICS if layer_key in t["applies_to"]]
        matrix_cols = ["entity_name", "ticker"] + [c for c in layer_topics if c in df.columns]
        layer_df = df[df["layer"] == layer_key].copy()
        if layer_df.empty:
            continue
        ws_matrix = wb.create_sheet(f"Raw Signal Matrix - {layer_label}")
        _write_matrix_sheet(ws_matrix, layer_df, matrix_cols, use_topic_display_names=True)

    # ── Methodology ─────────────────────────────────────────────────────────
    ws_method = wb.create_sheet("Methodology")
    _write_methodology_sheet(ws_method)

    EXCEL_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(EXCEL_OUTPUT))
    logger.info("Excel report saved to %s", EXCEL_OUTPUT)
    return EXCEL_OUTPUT


def _write_ranked_sheet(
    ws: Any,
    df: pd.DataFrame,
    columns: list[str],
    score_col: str,
) -> None:
    available_cols = [c for c in columns if c in df.columns]
    ws.append(["Rank"] + [c.replace("_", " ").title() for c in available_cols])
    _style_header(ws)

    for rank, (_, row) in enumerate(df.iterrows(), 1):
        values = [rank] + [row.get(c) for c in available_cols]
        ws.append(values)

    n_rows = len(df)
    n_cols = len(available_cols) + 1
    _style_body(ws, n_rows, n_cols)

    score_col_idx = available_cols.index(score_col) + 2 if score_col in available_cols else None
    if score_col_idx and n_rows > 1:
        col_letter = get_column_letter(score_col_idx)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{n_rows + 1}",
            ColorScaleRule(
                start_type="min", start_color="FF6B6B",
                mid_type="percentile", mid_value=50, mid_color="FFD93D",
                end_type="max", end_color="6BCB77",
            ),
        )

    _auto_width(ws)


def _write_matrix_sheet(
    ws: Any,
    df: pd.DataFrame,
    columns: list[str],
    use_topic_display_names: bool = False,
) -> None:
    available_cols = [c for c in columns if c in df.columns]
    headers = (
        [_topic_display_name(c) for c in available_cols]
        if use_topic_display_names
        else [c.replace("_", " ").title() for c in available_cols]
    )
    ws.append(headers)
    _style_header(ws)

    for _, row in df.iterrows():
        ws.append([row.get(c) for c in available_cols])

    n_rows = len(df)
    n_cols = len(available_cols)
    _style_body(ws, n_rows, n_cols)

    topic_start_col = 3
    if n_rows > 0 and n_cols >= topic_start_col:
        for col_idx in range(topic_start_col, n_cols + 1):
            col_letter = get_column_letter(col_idx)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{n_rows + 1}",
                ColorScaleRule(
                    start_type="min", start_color="E8F5E9",
                    end_type="max", end_color="2E7D32",
                ),
            )

    _auto_width(ws)


def _write_methodology_sheet(ws: Any) -> None:
    methodology_text = [
        ["Private Credit Stress Analyzer — Methodology"],
        [""],
        ["Data Source:"],
        ["  Bigdata API semantic search across news, filings, and transcripts."],
        [""],
        ["Entity Layers:"],
        ["  1. Lenders — BDCs, private credit funds, alternative asset managers"],
        ["  2. Borrowers — PE-backed companies with leveraged loan exposure"],
        ["  3. Banks — Back-leverage providers to private credit funds"],
        [""],
        ["Search Topics:"],
        ["  Each entity is searched against topic-specific query texts."],
        ["  Topics have polarity: 'positive' (strength) or 'negative' (stress)."],
        ["  {company} placeholder is replaced with entity name at runtime."],
        [""],
        ["Scoring Formula:"],
        ["  terms_power_score = positive_count / (positive_count + negative_count + 1) × 100"],
        ["  stress_score = 100 - terms_power_score"],
        [""],
        ["Layer-Specific Ranking:"],
        ["  Lenders: ranked by terms_power_score (high = strong)"],
        ["  Borrowers: ranked by stress_score (high = distressed)"],
        ["  Banks: ranked by net_position = market_share_gain − credit_pullback"],
        [""],
        ["Limitations:"],
        ["  Scores reflect mention frequency, not sentiment magnitude."],
        ["  Results depend on Bigdata index coverage and recency."],
    ]
    for row in methodology_text:
        ws.append(row)

    ws.column_dimensions["A"].width = 80
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.font = Font(color="000000", size=11)
            cell.fill = BODY_FILL
    ws["A1"].font = Font(bold=True, color="000000", size=14)


def _load_audit_docs(layer: str) -> list[dict[str, Any]]:
    """Load raw document chunks for the audit tab, limited to 5 per (entity, topic)."""
    from config.entities import ALL_ENTITIES
    from src.utils import sanitize_filename

    raw_dir = OUTPUT_DIR / "raw"
    layer_entities = [e for e in ALL_ENTITIES if e["layer"] == layer]
    layer_topics = [t for t in TOPICS if layer in t["applies_to"]]
    docs: list[dict[str, Any]] = []

    for entity in layer_entities:
        entity_slug = sanitize_filename(str(entity["name"]))
        for topic in layer_topics:
            topic_slug = sanitize_filename(str(topic["topic_name"]))
            path = raw_dir / f"{entity_slug}_{topic_slug}.json"
            if not path.exists():
                continue
            raw = json.loads(path.read_text())
            for result in raw.get("results", [])[:5]:
                content = result.get("content", "")
                docs.append({
                    "entity": entity["name"],
                    "topic": str(topic["topic_name"]).replace("_", " ").title(),
                    "polarity": topic["polarity"],
                    "headline": result.get("headline", ""),
                    "snippet": content[:250] + ("..." if len(content) > 250 else ""),
                    "timestamp": (result.get("timestamp") or "")[:10],
                    "url": result.get("url", ""),
                })
    return docs


def _prepare_layer_data(
    df: pd.DataFrame,
    layer: str,
) -> dict[str, Any]:
    """Prepare all chart/heatmap/theme data for a single layer."""
    layer_df = df[df["layer"] == layer].copy()
    layer_topics = [t for t in TOPICS if layer in t["applies_to"]]
    topic_names = [str(t["topic_name"]) for t in layer_topics]

    if layer == "lender":
        score_col = "terms_power_score"
        layer_df = layer_df.sort_values(score_col, ascending=True)
    elif layer == "borrower":
        score_col = "stress_score"
        layer_df = layer_df.sort_values(score_col, ascending=False)
    else:
        gain = layer_df["bank_market_share_gain"].fillna(0) if "bank_market_share_gain" in layer_df.columns else 0
        pullback = layer_df["bank_credit_pullback"].fillna(0) if "bank_credit_pullback" in layer_df.columns else 0
        layer_df["net_position"] = gain - pullback
        score_col = "net_position"
        layer_df = layer_df.sort_values(score_col, ascending=True)

    labels = layer_df["entity_name"].tolist()
    scores = layer_df[score_col].tolist()

    heatmap_rows: list[list[int]] = []
    available_topics = [t for t in topic_names if t in layer_df.columns]
    for _, row in layer_df.iterrows():
        heatmap_rows.append([
            int(v) if pd.notna(v) else 0
            for t in available_topics
            for v in [row.get(t, 0)]
        ])

    theme_topics: list[dict[str, Any]] = []
    for topic in layer_topics:
        tname = str(topic["topic_name"])
        total = int(layer_df[tname].fillna(0).sum()) if tname in layer_df.columns else 0
        theme_topics.append({
            "name": tname.replace("_", " ").title(),
            "polarity": topic["polarity"],
            "query": str(topic["topic_text"]),
            "count": total,
        })
    theme_topics.sort(key=lambda x: x["count"], reverse=True)

    audit_docs = _load_audit_docs(layer)

    out: dict[str, Any] = {
        "labels": labels,
        "scores": scores,
        "score_col": score_col,
        "heatmap_entities": labels,
        "heatmap_topics": [t.replace("_", " ").title() for t in available_topics],
        "heatmap_data": heatmap_rows,
        "theme_topics": theme_topics,
        "audit_docs": audit_docs,
        "entity_count": len(labels),
        "topic_count": len(layer_topics),
    }

    # Borrower radar: build from same heatmap data so radar and heatmap never diverge
    if layer == "borrower":
        radar_colors = ["#FF6B6B", "#FFD93D", "#4ECDC4", "#45B7D1", "#96CEB4", "#FFEAA7"]
        negative_topics = [
            str(t["topic_name"])
            for t in layer_topics
            if t["polarity"] == "negative" and str(t["topic_name"]) in available_topics
        ]
        radar_datasets = []
        for i, topic in enumerate(negative_topics):
            idx = available_topics.index(topic)
            radar_datasets.append({
                "label": topic.replace("_", " ").title(),
                "data": [heatmap_rows[row_idx][idx] for row_idx in range(len(heatmap_rows))],
                "borderColor": radar_colors[i % len(radar_colors)],
                "backgroundColor": radar_colors[i % len(radar_colors)] + "33",
            })
        out["radar_datasets"] = radar_datasets

    return out


def generate_html_dashboard(df: pd.DataFrame) -> Path:
    """Generate a standalone HTML dashboard with Chart.js visualizations."""
    lender_data = _prepare_layer_data(df, "lender")
    borrower_data = _prepare_layer_data(df, "borrower")
    bank_data = _prepare_layer_data(df, "bank")

    # Radar uses same entity order and counts as heatmap (from borrower_data)
    borrower_radar_labels = json.dumps(borrower_data["labels"])
    borrower_radar_datasets = json.dumps(borrower_data["radar_datasets"])

    html = _build_html(
        lender_data=lender_data,
        borrower_data=borrower_data,
        bank_data=bank_data,
        borrower_radar_labels=borrower_radar_labels,
        borrower_radar_datasets=borrower_radar_datasets,
    )

    HTML_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    HTML_OUTPUT.write_text(html)
    logger.info("HTML dashboard saved to %s", HTML_OUTPUT)
    return HTML_OUTPUT


def _themes_html(topics: list[dict[str, Any]]) -> str:
    """Build the Key Themes section with polarity, count, and query template."""
    items = ""
    for t in topics:
        pol_class = "positive" if t["polarity"] == "positive" else "negative"
        pol_icon = "+" if t["polarity"] == "positive" else "&minus;"
        query_escaped = t["query"].replace("{company}", "<em>{company}</em>")
        items += (
            f'<div class="theme-card {pol_class}">'
            f'<div class="theme-header">'
            f'<span class="theme-name">{t["name"]}</span>'
            f'<span class="pol-badge {pol_class[:3]}">{pol_icon} {t["polarity"].title()}</span>'
            f'<span class="theme-count">{t["count"]}</span>'
            f'</div>'
            f'<div class="theme-query">{query_escaped}</div>'
            f'</div>'
        )
    return items


def _build_html(
    lender_data: dict[str, Any],
    borrower_data: dict[str, Any],
    bank_data: dict[str, Any],
    borrower_radar_labels: str,
    borrower_radar_datasets: str,
) -> str:
    lender_themes = _themes_html(lender_data["theme_topics"])
    borrower_themes = _themes_html(borrower_data["theme_topics"])
    bank_themes = _themes_html(bank_data["theme_topics"])

    favicon_svg = "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 159.7 159.7'%3E%3Cpath fill='%234CA7F9' d='M38.11,0h83.48c21.03,0,38.11,17.08,38.11,38.11v83.48c0,21.03-17.08,38.11-38.11,38.11H38.11c-21.03,0-38.11-17.08-38.11-38.11V38.11C0,17.08,17.08,0,38.11,0Z'/%3E%3Cpath fill='%23FFFDF5' d='M105.69,137.06c-8.4,0-16.35-3.27-22.4-9.21-6.07-5.96-9.41-13.84-9.41-22.18v-51.63c0-11.38-8.72-20.3-19.87-20.3-5.4,0-10.44,2.12-14.21,5.96-3.74,3.82-5.81,8.91-5.81,14.34s2.06,10.52,5.81,14.34c3.76,3.84,8.81,5.96,14.21,5.96h13.36v11.09h-15.49c-8.26,0-15.83-3.26-21.32-9.19-5.4-5.83-8.37-13.71-8.37-22.2s3.34-16.22,9.41-22.18c6.05-5.94,14-9.21,22.4-9.21s16.36,3.27,22.4,9.21c6.07,5.96,9.41,13.84,9.41,22.18v51.63c0,11.38,8.72,20.3,19.86,20.3,5.4,0,10.44-2.12,14.21-5.96,3.74-3.82,5.81-8.91,5.81-14.34s-2.06-10.52-5.81-14.34c-3.76-3.84-8.81-5.96-14.21-5.96h-13.26v-11.09h15.4c8.26,0,15.83,3.26,21.32,9.19,5.4,5.82,8.37,13.71,8.37,22.2s-3.34,16.22-9.41,22.18c-6.05,5.94-14,9.21-22.4,9.21Z'/%3E%3C/svg%3E"

    # Per-layer descriptions for the 4 stat tiles (Entities, Topics, Highest, Lowest)
    _stat_descriptions = {
        "lender": (
            "BDCs and private credit managers in the universe.",
            "Strength and stress topics used for semantic search.",
            "Best terms power score across lenders (0–100).",
            "Lowest terms power; most stress signals.",
        ),
        "borrower": (
            "PE-backed companies with leveraged loan exposure.",
            "Resilience and distress topics per borrower.",
            "Most distressed borrower (highest stress score).",
            "Least distressed in the cohort.",
        ),
        "bank": (
            "Banks providing back-leverage to private credit.",
            "Contagion and market-share signal topics.",
            "Best net position (gain vs pullback).",
            "Worst net position; most pullback signal.",
        ),
    }

    _chart_descriptions = {
        "lender": (
            "Terms Power Score (0–100) = positive_mentions / (positive + negative + 1) × 100. "
            "Higher = stronger lender signal (spread power, fundraise resilience, NAV stability, covenant tightening "
            "vs redemption pressure, markdowns, waivers, PIK stress). Bars ranked by score descending."
        ),
        "borrower_radar": (
            "Each line is one distress topic (AI disruption, maturity wall, default risk, customer churn). "
            "Value at each axis = mention count for that borrower × topic. Same raw counts as the Signal Heatmap; only negative topics shown."
        ),
        "borrower_stress": (
            "Stress score (0–100) = 100 − terms_power_score, where terms_power = positive / (positive + negative + 1) × 100. "
            "High score = more distress signal relative to resilience (revenue growth, refinancing). Ranked by stress descending."
        ),
        "bank": (
            "Net position = market_share_gain_mentions − credit_pullback_mentions (and related negative topics). "
            "Positive = bank gaining share or conservative exposure; negative = pullback or contagion risk. Bars ranked by net position descending."
        ),
    }
    _heatmap_descriptions = {
        "lender": (
            "Each cell = count of search results that mention the lender and the topic (entity name must appear in headline or content). "
            "Rows = entities, columns = signal topics. These counts feed the Terms Power Score formula."
        ),
        "borrower": (
            "Each cell = mention count for that borrower × topic. Columns include positive (revenue growth, refinancing) and negative topics (AI disruption, maturity wall, default risk, customer churn). "
            "The stress score is the ratio of positive vs negative, not the raw sum."
        ),
        "bank": (
            "Each cell = mention count for that bank × topic. Net position score is the delta between positive (e.g. market share gain) and negative (credit pullback, margin call, contagion) topic counts."
        ),
    }

    def _layer_page(
        layer: str,
        data: dict[str, Any],
        chart_label: str,
        tab1_name: str,
        method_html: str,
        chart_panel_html: str | None = None,
        heatmap_desc: str = "",
    ) -> str:
        scores = data["scores"]
        hi = max(scores) if scores else 0
        lo = min(scores) if scores else 0
        hi_label = "Highest Score" if layer == "lender" else "Highest Stress" if layer == "borrower" else "Highest Net Pos."
        lo_label = "Lowest Score" if layer == "lender" else "Lowest Stress" if layer == "borrower" else "Lowest Net Pos."
        descs = _stat_descriptions[layer]
        active = ' active' if layer == 'lender' else ''
        if chart_panel_html is None:
            chart_panel_html = f'<div class="card"><h3>{chart_label}</h3><p class="card-desc">{_chart_descriptions.get(layer, "")}</p><canvas id="{layer}Chart"></canvas></div>'
        heatmap_block = f'<p class="card-desc">{heatmap_desc}</p>' if heatmap_desc else ''
        return f"""
    <div class="layer-page{active}" id="page-{layer}">
      <div class="stats">
        <div class="stat-card"><div class="stat-val">{data["entity_count"]}</div><div class="stat-label">Entities Tracked</div><div class="stat-desc">{descs[0]}</div></div>
        <div class="stat-card"><div class="stat-val">{data["topic_count"]}</div><div class="stat-label">Signal Topics</div><div class="stat-desc">{descs[1]}</div></div>
        <div class="stat-card"><div class="stat-val">{hi:.1f}</div><div class="stat-label">{hi_label}</div><div class="stat-desc">{descs[2]}</div></div>
        <div class="stat-card"><div class="stat-val">{lo:.1f}</div><div class="stat-label">{lo_label}</div><div class="stat-desc">{descs[3]}</div></div>
      </div>
      <div class="tabs" data-tabgroup="{layer}">
        <div class="tab active" onclick="switchTab('{layer}','chart')">{tab1_name}</div>
        <div class="tab" onclick="switchTab('{layer}','heatmap')">Signal Heatmap</div>
        <div class="tab" onclick="switchTab('{layer}','themes')">Key Themes</div>
        <div class="tab" onclick="switchTab('{layer}','audit')">Audit</div>
        <div class="tab" onclick="switchTab('{layer}','method')">Methodology</div>
      </div>
      <div class="tab-panel active" id="{layer}-chart">{chart_panel_html}</div>
      <div class="tab-panel" id="{layer}-heatmap"><div class="card"><h3>Signal Matrix</h3>{heatmap_block}<div class="heatmap-wrap" id="{layer}Heatmap"></div></div></div>
      <div class="tab-panel" id="{layer}-themes"><div class="card"><h3>Signal Topics &amp; Queries</h3><div class="themes-list">{_themes_html(data["theme_topics"])}</div></div></div>
      <div class="tab-panel" id="{layer}-audit"><div class="card"><h3>Document Audit</h3><div class="audit-filters" id="{layer}AuditFilters"></div><div class="audit-count" id="{layer}AuditCount"></div><div class="audit-table-wrap" id="{layer}AuditTable"></div></div></div>
      <div class="tab-panel" id="{layer}-method"><div class="card"><h3>How It Works</h3><div class="method-block">{method_html}</div></div></div>
    </div>"""

    _cov = f"Search date filter (Bigdata API): <strong>{SEARCH_DATE_LABEL}</strong>. Sources: news, filings, transcripts indexed by Bigdata."
    lender_method = '<h4>Scoring Formula</h4><p><code>terms_power_score = positive_count / (positive_count + negative_count + 1) &times; 100</code></p><p>High score = strong lender position (pricing power, fundraise resilience, covenant tightening).</p><h4>Data Pipeline</h4><p>1. Each entity is searched against each topic via Bigdata semantic search.<br>2. Results are filtered for documents that actually mention the entity name.<br>3. Mention counts are split by polarity (positive/negative) and aggregated.<br>4. Lenders are ranked by Terms Power Score (descending).</p><h4>Coverage</h4><p>' + _cov + '</p>'
    borrower_method = '<h4>Scoring Formula</h4><p><code>stress_score = 100 &minus; terms_power_score</code></p><p>High stress score = distressed borrower (AI disruption, maturity wall, default risk, customer churn).</p><h4>Data Pipeline</h4><p>1. Each borrower is searched against resilience and distress topics.<br>2. Radar chart shows per-topic negative signal intensity across all borrowers.<br>3. Borrowers are ranked by Stress Score (descending).</p><h4>Coverage</h4><p>' + _cov + ' Focus on PE-backed leveraged companies with private credit exposure.</p>'
    bank_method = '<h4>Scoring Formula</h4><p><code>net_position = market_share_gain_mentions &minus; credit_pullback_mentions</code></p><p>Positive net position = bank gaining share / conservative exposure. Negative = pulling back or contagion risk.</p><h4>Data Pipeline</h4><p>1. Each bank is searched against back-leverage, contagion, and market share topics.<br>2. Net position = delta between positive and negative topic mentions.<br>3. Banks ranked by net position (descending).</p><h4>Coverage</h4><p>' + _cov + ' Focus on banks providing back-leverage to private credit funds.</p>'

    lender_page = _layer_page(
        "lender",
        lender_data,
        "Lender Terms Power Score",
        "Score Analysis",
        lender_method,
        heatmap_desc=_heatmap_descriptions["lender"],
    )
    borrower_chart_panel = (
        '<div class="card"><h3>Distress radar</h3><p class="card-desc">'
        + _chart_descriptions["borrower_radar"]
        + '</p><canvas id="borrowerChart"></canvas></div>'
        + '<div class="card"><h3>Stress score by company</h3><p class="card-desc">'
        + _chart_descriptions["borrower_stress"]
        + '</p><canvas id="borrowerScoreChart"></canvas></div>'
    )
    borrower_page = _layer_page(
        "borrower",
        borrower_data,
        "Distress",
        "Distress",
        borrower_method,
        chart_panel_html=borrower_chart_panel,
        heatmap_desc=_heatmap_descriptions["borrower"],
    )
    bank_page = _layer_page(
        "bank",
        bank_data,
        "Bank Net Position Score",
        "Contagion Score",
        bank_method,
        heatmap_desc=_heatmap_descriptions["bank"],
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Private Credit Stress Analyzer — Bigdata</title>
<link rel="icon" type="image/svg+xml" href="{favicon_svg}">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ background:#0d1117; color:#c9d1d9; font-family:'Inter',system-ui,-apple-system,sans-serif; display:flex; height:100vh; overflow:hidden; }}

.sidebar {{ width:220px; min-width:220px; background:#161b22; border-right:1px solid #30363d; display:flex; flex-direction:column; }}
.sidebar-logo {{ padding:1.1rem 1.2rem; display:flex; align-items:center; gap:0.6rem; border-bottom:1px solid #30363d; }}
.sidebar-logo svg {{ width:26px; height:26px; flex-shrink:0; }}
.sidebar-logo span {{ font-size:1.05rem; font-weight:700; color:#fff; }}
.sidebar-logo .dot {{ color:#8b949e; font-weight:400; font-size:0.82rem; }}
.sidebar-section {{ padding:1rem 0 0; }}
.sidebar-section-label {{ padding:0 1.2rem; font-size:0.62rem; text-transform:uppercase; letter-spacing:1.5px; color:#484f58; font-weight:600; margin-bottom:0.5rem; }}
.nav-item {{ display:flex; align-items:center; gap:0.7rem; padding:0.6rem 1.2rem; cursor:pointer; color:#8b949e; font-size:0.85rem; font-weight:500; transition:all 0.15s; border-left:3px solid transparent; }}
.nav-item:hover {{ background:#1c2128; color:#c9d1d9; }}
.nav-item.active {{ background:rgba(76,167,249,0.08); color:#4CA7F9; border-left-color:#4CA7F9; }}
.nav-item svg {{ width:18px; height:18px; flex-shrink:0; opacity:0.7; }}
.nav-item.active svg {{ opacity:1; }}
.sidebar-footer {{ margin-top:auto; padding:1rem 1.2rem; border-top:1px solid #30363d; font-size:0.68rem; color:#484f58; line-height:1.5; }}

.main {{ flex:1; overflow-y:auto; display:flex; flex-direction:column; }}
.topbar {{ padding:1rem 2rem; border-bottom:1px solid #30363d; display:flex; align-items:center; justify-content:space-between; background:#161b22; min-height:56px; }}
.topbar h1 {{ font-size:1.15rem; font-weight:600; color:#fff; }}
.topbar .badge {{ background:rgba(76,167,249,0.15); color:#4CA7F9; font-size:0.7rem; padding:0.2rem 0.6rem; border-radius:99px; font-weight:600; }}
.content {{ flex:1; padding:1.5rem 2rem 2rem; overflow-y:auto; }}

.tabs {{ display:flex; gap:0; border-bottom:1px solid #30363d; margin-bottom:1.5rem; }}
.tab {{ padding:0.65rem 1.2rem; font-size:0.8rem; font-weight:500; color:#8b949e; cursor:pointer; border-bottom:2px solid transparent; transition:all 0.15s; }}
.tab:hover {{ color:#c9d1d9; }}
.tab.active {{ color:#00d4aa; border-bottom-color:#00d4aa; }}
.tab-panel {{ display:none; }}
.tab-panel.active {{ display:block; }}

.card {{ background:#161b22; border:1px solid #30363d; border-radius:8px; padding:1.5rem; margin-bottom:1.5rem; }}
.card h3 {{ color:#00d4aa; font-size:1rem; margin-bottom:1rem; font-weight:600; }}
canvas {{ max-height:420px; }}

.heatmap-wrap {{ overflow-x:auto; }}
.hm {{ border-collapse:collapse; width:100%; font-size:0.75rem; }}
.hm th {{ background:#1a1a2e; color:#00d4aa; padding:8px 10px; text-align:center; font-weight:600; white-space:nowrap; position:sticky; top:0; z-index:1; }}
.hm td {{ padding:7px 10px; text-align:center; border:1px solid #21262d; }}
.hm tr:nth-child(even) {{ background:#0d1117; }}
.hm tr:nth-child(odd) {{ background:#161b22; }}

.themes-list {{ display:flex; flex-direction:column; gap:0.5rem; }}
.theme-card {{ padding:0.75rem 1rem; border-radius:6px; }}
.theme-card.positive {{ background:rgba(0,212,170,0.06); border-left:3px solid #00d4aa; }}
.theme-card.negative {{ background:rgba(255,107,107,0.06); border-left:3px solid #FF6B6B; }}
.theme-header {{ display:flex; align-items:center; gap:0.6rem; }}
.theme-name {{ font-size:0.85rem; font-weight:600; color:#c9d1d9; flex:1; }}
.theme-count {{ font-weight:700; font-size:1rem; min-width:2.5rem; text-align:right; }}
.theme-card.positive .theme-count {{ color:#00d4aa; }}
.theme-card.negative .theme-count {{ color:#FF6B6B; }}
.theme-query {{ font-size:0.75rem; color:#8b949e; font-style:italic; margin-top:0.35rem; padding-left:0.1rem; }}
.theme-query em {{ color:#c9d1d9; font-style:normal; font-weight:600; }}
.pol-badge {{ font-size:0.65rem; padding:0.1rem 0.45rem; border-radius:99px; font-weight:600; white-space:nowrap; }}
.pol-badge.pos {{ background:rgba(0,212,170,0.15); color:#00d4aa; }}
.pol-badge.neg {{ background:rgba(255,107,107,0.15); color:#FF6B6B; }}

.audit-filters {{ display:flex; gap:0.75rem; margin-bottom:1rem; flex-wrap:wrap; }}
.audit-filters select {{ background:#0d1117; color:#c9d1d9; border:1px solid #30363d; border-radius:6px; padding:0.4rem 0.7rem; font-size:0.8rem; cursor:pointer; min-width:160px; }}
.audit-filters select:focus {{ border-color:#00d4aa; outline:none; }}
.audit-count {{ font-size:0.75rem; color:#8b949e; margin-bottom:0.75rem; }}
.audit-table-wrap {{ overflow-x:auto; max-height:500px; overflow-y:auto; }}
.at {{ border-collapse:collapse; width:100%; font-size:0.78rem; }}
.at th {{ background:#1a1a2e; color:#00d4aa; padding:8px 10px; text-align:left; font-weight:600; white-space:nowrap; position:sticky; top:0; z-index:1; }}
.at td {{ padding:8px 10px; border-bottom:1px solid #21262d; vertical-align:top; }}
.at tr:hover {{ background:#1c2128; }}
.at .td-hl {{ font-weight:600; color:#c9d1d9; max-width:280px; }}
.at .td-snip {{ color:#8b949e; font-size:0.73rem; max-width:350px; line-height:1.4; }}
.at .td-link a {{ color:#4CA7F9; text-decoration:none; font-size:0.72rem; }}
.at .td-link a:hover {{ text-decoration:underline; }}
.at .td-date {{ white-space:nowrap; color:#8b949e; }}

.method-block {{ font-size:0.85rem; line-height:1.8; color:#8b949e; }}
.method-block h4 {{ color:#c9d1d9; font-size:0.9rem; margin:1.2rem 0 0.3rem; }}
.method-block code {{ background:#1a1a2e; padding:0.15rem 0.4rem; border-radius:4px; color:#00d4aa; font-size:0.8rem; }}

.stats {{ display:flex; gap:1rem; margin-bottom:1.5rem; }}
.stat-card {{ flex:1; background:#161b22; border:1px solid #30363d; border-radius:8px; padding:1rem 1.2rem; text-align:center; }}
.stat-val {{ font-size:1.6rem; font-weight:700; color:#00d4aa; }}
.stat-label {{ font-size:0.72rem; color:#8b949e; text-transform:uppercase; letter-spacing:0.5px; margin-top:0.2rem; }}
.stat-desc {{ font-size:0.7rem; color:#484f58; line-height:1.3; margin-top:0.35rem; max-width:140px; margin-left:auto; margin-right:auto; }}

.card-desc {{ font-size:0.78rem; color:#8b949e; margin-bottom:0.75rem; }}

.layer-page {{ display:none; }}
.layer-page.active {{ display:block; }}
</style>
</head>
<body>

<nav class="sidebar">
  <div class="sidebar-logo">
    <svg viewBox="0 0 159.7 159.7"><path fill="#4CA7F9" d="M38.11,0h83.48c21.03,0,38.11,17.08,38.11,38.11v83.48c0,21.03-17.08,38.11-38.11,38.11H38.11c-21.03,0-38.11-17.08-38.11-38.11V38.11C0,17.08,17.08,0,38.11,0Z"/><path fill="#FFFDF5" d="M105.69,137.06c-8.4,0-16.35-3.27-22.4-9.21-6.07-5.96-9.41-13.84-9.41-22.18v-51.63c0-11.38-8.72-20.3-19.87-20.3-5.4,0-10.44,2.12-14.21,5.96-3.74,3.82-5.81,8.91-5.81,14.34s2.06,10.52,5.81,14.34c3.76,3.84,8.81,5.96,14.21,5.96h13.36v11.09h-15.49c-8.26,0-15.83-3.26-21.32-9.19-5.4-5.83-8.37-13.71-8.37-22.2s3.34-16.22,9.41-22.18c6.05-5.94,14-9.21,22.4-9.21s16.36,3.27,22.4,9.21c6.07,5.96,9.41,13.84,9.41,22.18v51.63c0,11.38,8.72,20.3,19.86,20.3,5.4,0,10.44-2.12,14.21-5.96,3.74-3.82,5.81-8.91,5.81-14.34s-2.06-10.52-5.81-14.34c-3.76-3.84-8.81-5.96-14.21-5.96h-13.26v-11.09h15.4c8.26,0,15.83,3.26,21.32,9.19,5.4,5.82,8.37,13.71,8.37,22.2s-3.34,16.22-9.41,22.18c-6.05,5.94-14,9.21-22.4,9.21Z"/></svg>
    <span>bigdata<span class="dot">.com</span></span>
  </div>
  <div class="sidebar-section">
    <div class="sidebar-section-label">Analysis Layers</div>
    <div class="nav-item active" data-layer="lender" onclick="switchLayer('lender')">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="7" height="7" rx="1"/><rect x="14" y="3" width="7" height="7" rx="1"/><rect x="3" y="14" width="7" height="7" rx="1"/><rect x="14" y="14" width="7" height="7" rx="1"/></svg>
      Lenders
    </div>
    <div class="nav-item" data-layer="borrower" onclick="switchLayer('borrower')">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M12 2L2 7l10 5 10-5-10-5z"/><path d="M2 17l10 5 10-5"/><path d="M2 12l10 5 10-5"/></svg>
      Borrowers
    </div>
    <div class="nav-item" data-layer="bank" onclick="switchLayer('bank')">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M3 21h18M3 10h18M5 6l7-3 7 3M4 10v11M20 10v11M8 14v3M12 14v3M16 14v3"/></svg>
      Banks
    </div>
  </div>
  <div class="sidebar-footer">Private Credit Stress Analyzer<br>Powered by Bigdata API</div>
</nav>

<div class="main">
  <div class="topbar">
    <h1 id="pageTitle">Lender Terms Power Analysis</h1>
    <span class="badge" id="pageBadge">{lender_data["entity_count"]} entities &middot; {lender_data["topic_count"]} topics</span>
  </div>
  <div class="content">
    {lender_page}
    {borrower_page}
    {bank_page}
  </div>
</div>

<script>
Chart.defaults.color = '#c9d1d9';
Chart.defaults.borderColor = '#30363d';

const pageTitles = {{ lender:'Lender Terms Power Analysis', borrower:'Borrower Distress Analysis', bank:'Bank Contagion Analysis' }};
const pageBadges = {{ lender:'{lender_data["entity_count"]} entities &middot; {lender_data["topic_count"]} topics', borrower:'{borrower_data["entity_count"]} entities &middot; {borrower_data["topic_count"]} topics', bank:'{bank_data["entity_count"]} entities &middot; {bank_data["topic_count"]} topics' }};

const auditData = {{
  lender: {json.dumps(lender_data["audit_docs"], default=str)},
  borrower: {json.dumps(borrower_data["audit_docs"], default=str)},
  bank: {json.dumps(bank_data["audit_docs"], default=str)}
}};

let chartsInitialized = {{ lender:false, borrower:false, bank:false }};
let auditsInitialized = {{ lender:false, borrower:false, bank:false }};

function switchLayer(layer) {{
  document.querySelectorAll('.nav-item').forEach(n => n.classList.remove('active'));
  document.querySelector('.nav-item[data-layer="'+layer+'"]').classList.add('active');
  document.querySelectorAll('.layer-page').forEach(p => p.classList.remove('active'));
  document.getElementById('page-'+layer).classList.add('active');
  document.getElementById('pageTitle').textContent = pageTitles[layer];
  document.getElementById('pageBadge').innerHTML = pageBadges[layer];
  if (!chartsInitialized[layer]) initCharts(layer);
  if (!auditsInitialized[layer]) initAudit(layer);
}}

function switchTab(layer, tab) {{
  const page = document.getElementById('page-'+layer);
  page.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  page.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
  event.target.classList.add('active');
  document.getElementById(layer+'-'+tab).classList.add('active');
  if (tab === 'audit' && !auditsInitialized[layer]) initAudit(layer);
}}

function buildHeatmap(id, entities, topics, data) {{
  let mx = 1;
  data.forEach(r => r.forEach(v => {{ if(v>mx) mx=v; }}));
  let h = '<table class="hm"><thead><tr><th>Entity</th>';
  topics.forEach(t => {{ h += '<th>'+t+'</th>'; }});
  h += '</tr></thead><tbody>';
  data.forEach((row,i) => {{
    h += '<tr><td style="text-align:left;font-weight:600;white-space:nowrap;">'+entities[i]+'</td>';
    row.forEach(v => {{
      const p = Math.min(v/mx,1);
      h += '<td style="background:rgba('+Math.round(13+p*(0-13))+','+Math.round(17+p*(212-17))+','+Math.round(23+p*(170-23))+','+(0.15+p*0.85)+');color:#fff;">'+v+'</td>';
    }});
    h += '</tr>';
  }});
  h += '</tbody></table>';
  document.getElementById(id).innerHTML = h;
}}

function initAudit(layer) {{
  auditsInitialized[layer] = true;
  const docs = auditData[layer];
  if (!docs || !docs.length) return;

  const entities = [...new Set(docs.map(d => d.entity))].sort();
  const topics = [...new Set(docs.map(d => d.topic))].sort();
  const polarities = ['positive','negative'];

  const fc = document.getElementById(layer+'AuditFilters');
  fc.innerHTML = '<select id="'+layer+'FEntity"><option value="">All Entities</option>'+entities.map(e=>'<option>'+e+'</option>').join('')+'</select>'
    + '<select id="'+layer+'FTopic"><option value="">All Topics</option>'+topics.map(t=>'<option>'+t+'</option>').join('')+'</select>'
    + '<select id="'+layer+'FPol"><option value="">All Polarities</option>'+polarities.map(p=>'<option value="'+p+'">'+p[0].toUpperCase()+p.slice(1)+'</option>').join('')+'</select>';

  [layer+'FEntity',layer+'FTopic',layer+'FPol'].forEach(id => {{
    document.getElementById(id).addEventListener('change', () => renderAudit(layer));
  }});
  renderAudit(layer);
}}

function renderAudit(layer) {{
  const docs = auditData[layer];
  const ent = document.getElementById(layer+'FEntity').value;
  const top = document.getElementById(layer+'FTopic').value;
  const pol = document.getElementById(layer+'FPol').value;

  let filtered = docs;
  if (ent) filtered = filtered.filter(d => d.entity === ent);
  if (top) filtered = filtered.filter(d => d.topic === top);
  if (pol) filtered = filtered.filter(d => d.polarity === pol);

  document.getElementById(layer+'AuditCount').textContent = 'Showing '+filtered.length+' of '+docs.length+' documents';

  let h = '<table class="at"><thead><tr><th>Entity</th><th>Topic</th><th>Polarity</th><th>Headline</th><th>Content</th><th>Date</th><th>Source</th></tr></thead><tbody>';
  filtered.slice(0,200).forEach(d => {{
    const pc = d.polarity === 'positive' ? 'pos' : 'neg';
    const pi = d.polarity === 'positive' ? '+' : '&minus;';
    const link = d.url ? '<a href="'+d.url+'" target="_blank" rel="noopener">View &rarr;</a>' : '';
    h += '<tr><td style="font-weight:600;white-space:nowrap;">'+d.entity+'</td>'
      + '<td style="white-space:nowrap;">'+d.topic+'</td>'
      + '<td><span class="pol-badge '+pc+'">'+pi+' '+d.polarity+'</span></td>'
      + '<td class="td-hl">'+d.headline+'</td>'
      + '<td class="td-snip">'+d.snippet+'</td>'
      + '<td class="td-date">'+d.timestamp+'</td>'
      + '<td class="td-link">'+link+'</td></tr>';
  }});
  h += '</tbody></table>';
  document.getElementById(layer+'AuditTable').innerHTML = h;
}}

function initCharts(layer) {{
  chartsInitialized[layer] = true;
  if (layer === 'lender') {{
    const s = {json.dumps(lender_data["scores"])};
    new Chart(document.getElementById('lenderChart'), {{ type:'bar', data:{{ labels:{json.dumps(lender_data["labels"])}, datasets:[{{ label:'Terms Power Score', data:s, backgroundColor:s.map(v=>v>50?'#00d4aa':v>30?'#FFD93D':'#FF6B6B'), borderRadius:4 }}] }}, options:{{ indexAxis:'y', responsive:true, plugins:{{legend:{{display:false}}}}, scales:{{ x:{{min:0,max:100,grid:{{color:'#21262d'}}}}, y:{{grid:{{display:false}}}} }} }} }});
    buildHeatmap('lenderHeatmap', {json.dumps(lender_data["heatmap_entities"])}, {json.dumps(lender_data["heatmap_topics"])}, {json.dumps(lender_data["heatmap_data"])});
  }}
  if (layer === 'borrower') {{
    new Chart(document.getElementById('borrowerChart'), {{ type:'radar', data:{{ labels:{borrower_radar_labels}, datasets:{borrower_radar_datasets} }}, options:{{ responsive:true, plugins:{{legend:{{position:'bottom',labels:{{boxWidth:12}}}}}}, scales:{{ r:{{beginAtZero:true, grid:{{color:'#21262d'}}, angleLines:{{color:'#21262d'}}, pointLabels:{{font:{{size:10}},color:'#c9d1d9'}}}} }} }} }});
    const bScores = {json.dumps(borrower_data["scores"])};
    const bLabels = {json.dumps(borrower_data["labels"])};
    new Chart(document.getElementById('borrowerScoreChart'), {{ type:'bar', data:{{ labels:bLabels, datasets:[{{ label:'Stress Score', data:bScores, backgroundColor:bScores.map(v=>v>=70?'#FF6B6B':v>=50?'#FFD93D':'#00d4aa'), borderRadius:4 }}] }}, options:{{ indexAxis:'y', responsive:true, plugins:{{legend:{{display:false}}}}, scales:{{ x:{{min:0,max:100,grid:{{color:'#21262d'}}}}, y:{{grid:{{display:false}}}} }} }} }});
    buildHeatmap('borrowerHeatmap', {json.dumps(borrower_data["heatmap_entities"])}, {json.dumps(borrower_data["heatmap_topics"])}, {json.dumps(borrower_data["heatmap_data"])});
  }}
  if (layer === 'bank') {{
    const bs = {json.dumps(bank_data["scores"])};
    new Chart(document.getElementById('bankChart'), {{ type:'bar', data:{{ labels:{json.dumps(bank_data["labels"])}, datasets:[{{ label:'Net Position', data:bs, backgroundColor:bs.map(v=>v>0?'#00d4aa':'#FF6B6B'), borderRadius:4 }}] }}, options:{{ indexAxis:'y', responsive:true, plugins:{{legend:{{display:false}}}}, scales:{{ x:{{grid:{{color:'#21262d'}}}}, y:{{grid:{{display:false}}}} }} }} }});
    buildHeatmap('bankHeatmap', {json.dumps(bank_data["heatmap_entities"])}, {json.dumps(bank_data["heatmap_topics"])}, {json.dumps(bank_data["heatmap_data"])});
  }}
}}

initCharts('lender');
initAudit('lender');
</script>
</body>
</html>"""


def generate_reports(df: pd.DataFrame | None = None) -> None:
    """Generate both Excel and HTML reports from scores."""
    if df is None:
        if not SCORES_CSV.exists():
            raise FileNotFoundError(
                f"Scores file not found at {SCORES_CSV}. Run scorer first."
            )
        df = pd.read_csv(SCORES_CSV)

    console.rule("[bold cyan]Report Generation")
    excel_path = generate_excel(df)
    html_path = generate_html_dashboard(df)
    console.print(f"[green]Excel:[/green] {excel_path}")
    console.print(f"[green]HTML:[/green]  {html_path}")


if __name__ == "__main__":
    generate_reports()
